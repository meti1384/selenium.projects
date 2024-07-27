from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from anticaptchaofficial.recaptchav2proxyless import *
from ruamel.std.zipfile import delete_from_zip_file
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium import webdriver
import mysql.connector
from time import sleep
import openpyxl
import zipfile
import shutil
import os



DOWNLOAD_DEAULAT_DIRECTORY=r""   # => this directory must be empty
options=Options()
options.add_experimental_option("prefs", {"download.default_directory": DOWNLOAD_DEAULAT_DIRECTORY})
service=Service()
driver=webdriver.Edge(service=service, options=options)
driver.implicitly_wait(20)
action=ActionChains(driver)


def get_url(url):
    try:
        driver.get(url)
        result=driver.execute_script("return document.readyState")
        if result!="complete":
            raise Exception("site is not loaded!")
    except:
        sleep(12)

    


def extract_data_form_excel(path):
    file_path = path

    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active

    data = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        data.append(row)

    workbook.close()
    return data


def captchat_solve(url, website_key):
    solver = recaptchaV2Proxyless()
    solver.set_verbose(1)
    solver.set_key("")
    solver.set_website_url(url)
    solver.set_website_key(website_key)

    g_response = solver.solve_and_return_solution()
    if g_response!= 0:
        print("g_response"+g_response)
    else:
        print("task finished with error"+solver.error_code)


    driver.execute_script('document.getElementById("g-recaptcha-response").innerHTML = "<g_response>";')
    sleep(2)
    driver.execute_script(f"___grecaptcha_cfg.clients[0].S.S.callback('{g_response}');")
    sleep(15)




def submit_login_data():
    driver.find_element(By.CSS_SELECTOR , "input#username").send_keys("phone_number")
    driver.find_element(By.CSS_SELECTOR, "button[class^=btn]").click()
    driver.find_element(By.CSS_SELECTOR , "input#password").send_keys("password")
    driver.find_element(By.XPATH, "//button[@class='btn btn--orange btn--full' and @type='submit']").click()

    

def login(url, website_key):
            
           
                 
                try:
                    try:
                        driver.find_element(By.CSS_SELECTOR , "input#username")
                    except:
                        return
                      
                    driver.find_element(By.XPATH, "//iframe[@title='reCAPTCHA']")
                except:
                    submit_login_data()
                    return
                
                try:
                    captchat_solve(url, website_key)
                    submit_login_data()
                    sleep(4)
                    return

                except:

                    try:
                        driver.refresh()
                        captchat_solve(url, website_key)
                        submit_login_data()
                    
                    except:
                        
                            if driver.find_element(By.XPATH, "#__next > div.page-sign > div > div > div > div > div > div.form__wrap > div > div]"):
                                raise Exception("Maby captcha service have mistake! please run program again!")
                              
                            raise ("Maby the callback path expired(for fix this go to the this site https://discourse.openbullet.dev/t/guide-recaptcha-v2-bypass-and-callbacks/3252)") 
            



def download_old_plugin_version():
    element=driver.find_element(By.XPATH, '//a[text()="دانلود"]')
    element.click()
    sleep(10)



def find_upload_element():
    element=driver.find_element(By.CSS_SELECTOR, "#__next > div > div > main > div.add-product > form > div.add-product__body > div.max-lg > div > div.col-xs-12.col-lg-5 > div > div > div:nth-child(2) > div > div > div.form__root > div > div > div > div")
    action.move_to_element(element).click().perform()
    file_input=driver.find_element(By.CSS_SELECTOR, "#__next > div > div > main > div.add-product > form > div.add-product__body > div.max-lg > div > div.col-xs-12.col-lg-5 > div > div > div:nth-child(2) > div > div > div:nth-child(1) > div > div > div.uppy-Dashboard-inner > div > div.uppy-Dashboard-AddFiles > input:nth-child(1)")
    return file_input

 

def enter_new_version(version):
        enter_version_input=driver.find_element(By.XPATH, "//input[@name='version']")
        action.move_to_element(enter_version_input).click().key_down(Keys.CONTROL).send_keys("a").key_up(Keys.CONTROL).send_keys(Keys.DELETE ).perform()
        sleep(4)
        enter_version_input.send_keys(version)

        sleep(6)
        return


def save_product():
    save_product_button=driver.find_element(By.XPATH, "//button[text()='ذخیره محصول']")
    action.move_to_element(save_product_button).click().perform()
    WebDriverWait(driver, 50).until(
         EC.presence_of_element_located((By.XPATH, "//*[@class='main-alerts']"))
    )

    


list_of_urles=[]
def check_update_notification(status, url):
    checkbox_element=driver.find_element(By.CSS_SELECTOR, "#update_notify")
    action.move_to_element(checkbox_element)
    value_of_checkbox=checkbox_element.get_property("value")
    if value_of_checkbox == "false":
        if status == 1:
            checkbox_element.click()
            list_of_urles.append(url)
            
        accept_zhaket_rules=driver.find_element(By.XPATH, "//input[@name='terms_acceptance']")
        if accept_zhaket_rules.get_property("vlaue")=="false":
            accept_zhaket_rules.click()
                      
    save_product()
    sleep(5)




def delete_files(path_extract_file, path_of_file):
        shutil.rmtree(os.path.join(path_extract_file, "1- plugin"))
        for item_path in os.listdir(DOWNLOAD_DEAULAT_DIRECTORY):
            old_zip_path=os.path.join(DOWNLOAD_DEAULAT_DIRECTORY, item_path)
            delete_from_zip_file(path_of_file, pattern='.*.zip')
            return old_zip_path


def put_old_zip_in_main_zip(old_zip_path, old_version, zip_path, name, ):
     with zipfile.ZipFile(zip_path, "a") as zip_ref:
                arcname = f"1- plugin\{name}_{old_version}.zip"
                zip_ref.write(old_zip_path, arcname=arcname)
                os.remove(old_zip_path)
            


def upload_new_plugin_and_save_old_plugin(path_of_file, path_extract_file, name, version, status, url):
    with zipfile.ZipFile(path_of_file, "r") as zip_ref:  
        for item in (zip_ref.namelist()):
            if item.startswith("1- plugin") and item.endswith(".zip"):  
                zip_ref.extract(item, path=path_extract_file)
                new_zip_path=os.path.join(path_extract_file,item)

                download_old_plugin_version()   
                find_upload_element().send_keys(new_zip_path)
                WebDriverWait(driver, 50).until(
                    EC.text_to_be_present_in_element_attribute((By.XPATH, 
                                                            '//*[@class="uppy-StatusBar-content"]'), 'title', 'کامل شد'))
                
                sleep(3)
                save_old_version_value=driver.find_element(By.XPATH, "//input[@name='version']").get_attribute('value')
                enter_new_version(version)
                check_update_notification(status, url)
                put_old_zip_in_main_zip(delete_files(path_extract_file, path_of_file), save_old_version_value,path_of_file, name)
                return



def remove_row_in_excel_file(path):
    
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook["Sheet1"]

    for i in range(1,6):
        worksheet.cell(row=2, column=i).value=None
    worksheet.delete_rows(2)

    workbook.save(path)
    
    
def get_plugins_url_which_insert_to_database(list_of_url):
    connect_database=mysql.connector.connect(
          host="",
          user="",
          password="",
          database=""
     )
    cursor=connect_database.cursor()
    for url in list_of_url:
        insert_url_in_table="""  
                    insert into tablename(column_name)
                    values(%s);"""
    
        cursor.execute(insert_url_in_table, (url,))



    connect_database.commit()
    cursor.close()
    connect_database.close()